"""
excel_writer.py – Zápis článků do výstupního Excelu.

Struktura výstupního souboru:
  - List "PŘEHLED"  – souhrnný list s checkboxy pro kolegyně
  - Jeden list per web (pojmenovaný podle source_name)

Checkboxy: Sloupec A obsahuje symboly ☐ (nezpracováno) / ☑ (zpracováno).
Kolegyně klikne na buňku, z dropdown vybere ☑ a řádek se automaticky
obarví zeleně (podmíněné formátování). Funguje v Excelu i SharePointu.
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "vysledky.xlsx")

# ── Symboly checkboxu ─────────────────────────────────────────────────────────
CHECKBOX_UNCHECKED = "☐"   # Nezpracováno (default)
CHECKBOX_CHECKED   = "☑"   # Zpracováno

# ── Barvy ─────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"   # Tmavě modrá – záhlaví
COLOR_HEADER_FONT = "FFFFFF"   # Bílá – text záhlaví
COLOR_ROW_ODD     = "DEEAF1"   # Světle modrá – liché řádky
COLOR_ROW_EVEN    = "FFFFFF"   # Bílá – sudé řádky
COLOR_DONE        = "C6EFCE"   # Světle zelená – zpracováno (☑)
COLOR_DONE_FONT   = "276221"   # Tmavě zelená – font zpracováno
COLOR_OSTRAVA_BG  = "FFF2CC"   # Žlutá – Ostrava řádky (nezpracované)
COLOR_ACCENT      = "2E75B6"   # Modrá – hyperlinky

HEADER_HEIGHT = 22
ROW_HEIGHT    = 18

# ── Definice sloupců ──────────────────────────────────────────────────────────
PREHLED_COLUMNS = [
    ("✓", 5),               # A – checkbox
    ("Web", 18),
    ("Titulek", 55),
    ("URL", 50),
    ("Datum publikace", 18),
    ("Město", 12),
    ("Datum přidání", 18),
    ("Autor", 20),
    ("Kategorie", 22),
]

DETAIL_COLUMNS = [
    ("Titulek", 55),
    ("URL", 50),
    ("Datum publikace", 18),
    ("Perex", 70),
    ("Autor", 20),
    ("Kategorie", 22),
    ("Město", 12),
    ("Datum scrapování", 18),
]

NUM_PREHLED_COLS = len(PREHLED_COLUMNS)


def _style_header_row(ws, columns: list):
    """Nastyluje řádek záhlaví a nastaví šířky sloupců."""
    ws.row_dimensions[1].height = HEADER_HEIGHT
    hdr_font = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=10)
    hdr_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"


def _style_data_row(ws, row_num: int, num_cols: int, is_ostrava: bool = False):
    """Nastyluje datový řádek (výchozí barvy – přepíše podmíněné formátování)."""
    ws.row_dimensions[row_num].height = ROW_HEIGHT
    if is_ostrava:
        bg = COLOR_OSTRAVA_BG
    else:
        bg = COLOR_ROW_ODD if row_num % 2 == 1 else COLOR_ROW_EVEN
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name="Calibri", size=9)
    align = Alignment(vertical="center", wrap_text=False)

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _add_checkbox_column(ws, data_start_row: int, max_rows: int = 5000):
    """
    Nastaví sloupec A pro rychlé označení 'x' (nebo jakýkoliv znak).
    Funguje bleskově pro 1 řádek (napsat x + Enter) i pro hromadný výběr
    více řádků (označit myší -> napsat x -> Ctrl+Enter).
    Jakýkoliv zápis do sloupce A obarví celý řádek zeleně.
    """
    last_row = data_start_row + max_rows

    # Podmíněné formátování: když sloupec A není prázdný a není ☐ -> celý řádek zelený
    done_fill = PatternFill("solid", fgColor=COLOR_DONE)
    done_font = Font(name="Calibri", size=9, color=COLOR_DONE_FONT, bold=False)

    col_last = get_column_letter(NUM_PREHLED_COLS)
    cf_range = f"A{data_start_row}:{col_last}{last_row}"

    # Formule: $A2<>"" AND $A2<>"☐" → celý řádek zelený
    rule = FormulaRule(
        formula=[f'AND($A{data_start_row}<>"", $A{data_start_row}<>"☐")'],
        fill=done_fill,
        font=done_font,
    )
    ws.conditional_formatting.add(cf_range, rule)


def _get_or_create_sheet(wb: Workbook, name: str) -> tuple:
    """Vrátí existující nebo nový list a příznak jestli je nový."""
    safe_name = name[:31]
    if safe_name in wb.sheetnames:
        return wb[safe_name], False
    ws = wb.create_sheet(title=safe_name)
    return ws, True


def write_articles_to_excel(new_articles: list[dict], output_path: str = None) -> int:
    """
    Zapíše nové články na začátek (HODNĚ/NAHOŘE, pod hlavičku) výstupního Excelu.
    Vrátí počet zapsaných článků.
    """
    if not new_articles:
        return 0

    path = output_path or OUTPUT_FILE

    # Vytvořit složku pokud neexistuje
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)

    # Načíst existující soubor nebo vytvořit nový
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Souhrnný list PŘEHLED ────────────────────────────────────────────────
    is_new_prehled = "PŘEHLED" not in wb.sheetnames
    if is_new_prehled:
        ws_prehled = wb.create_sheet(title="PŘEHLED", index=0)
        _style_header_row(ws_prehled, PREHLED_COLUMNS)
    else:
        ws_prehled = wb["PŘEHLED"]

    # Vložit řádky na začátek (pod hlavičku na řádek 2)
    if not is_new_prehled and ws_prehled.max_row > 1:
        ws_prehled.insert_rows(2, amount=len(new_articles))

    # ── Zápis článků do PŘEHLEDU (od nejnovějšího po nejstarší na řádek 2+) ──
    for idx, article in enumerate(new_articles):
        source     = article.get("source_name", "unknown")
        title      = article.get("title", "")
        url        = article.get("url", "")
        date       = article.get("date", "")
        city       = article.get("city", "Jiné")
        author     = article.get("author", "")
        cat        = article.get("category", "")
        perex      = article.get("perex", "")
        scraped    = article.get("scraped_at", "")
        is_ostrava = (city == "Ostrava")

        # ── Detail list (per web) ──────────────────────────────────────────
        ws_detail, is_new_detail = _get_or_create_sheet(wb, source)
        if is_new_detail:
            _style_header_row(ws_detail, DETAIL_COLUMNS)
            detail_row = 2
        else:
            ws_detail.insert_rows(2, amount=1)
            detail_row = 2

        _style_data_row(ws_detail, detail_row, len(DETAIL_COLUMNS), is_ostrava)
        for col_i, val in enumerate([title, url, date, perex, author, cat, city, scraped], start=1):
            cell = ws_detail.cell(row=detail_row, column=col_i, value=val)
            if col_i == 2 and url:
                cell.hyperlink = url
                cell.font = Font(name="Calibri", size=9, color=COLOR_ACCENT, underline="single")

        # ── Souhrnný list PŘEHLED (řádek 2 + idx) ──────────────────────────
        prehled_row = 2 + idx
        _style_data_row(ws_prehled, prehled_row, NUM_PREHLED_COLS, is_ostrava)

        # Sloupec A = ☐ (nezpracováno)
        cb_cell = ws_prehled.cell(row=prehled_row, column=1, value=CHECKBOX_UNCHECKED)
        cb_cell.alignment = Alignment(horizontal="center", vertical="center")
        cb_cell.font = Font(name="Segoe UI Symbol", size=12, bold=False)

        prehled_data = [source, title, url, date, city, scraped, author, cat]
        for col_i, val in enumerate(prehled_data, start=2):
            cell = ws_prehled.cell(row=prehled_row, column=col_i, value=val)
            if col_i == 4 and url:   # sloupec D = URL
                cell.hyperlink = url
                cell.font = Font(name="Calibri", size=9, color=COLOR_ACCENT, underline="single")

    # ── Checkbox dropdown + podmíněné formátování ────────────────────────────
    _add_checkbox_column(ws_prehled, data_start_row=2, max_rows=5000)

    # PŘEHLED vždy jako první list
    if "PŘEHLED" in wb.sheetnames:
        p_idx = wb.sheetnames.index("PŘEHLED")
        if p_idx != 0:
            wb.move_sheet("PŘEHLED", offset=-p_idx)

    wb.save(path)
    return len(new_articles)
