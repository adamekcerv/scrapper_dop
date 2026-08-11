import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PŘEHLED"

ws.append(["Hotovo (x)", "Web", "Titulek"])
ws.append(["", "polar.cz", "Článek 1 - nepřečteno"])
ws.append(["x", "denik.cz", "Článek 2 - přečteno x"])
ws.append(["1", "drbna.cz", "Článek 3 - přečteno 1"])
ws.append(["ano", "okraj.cz", "Článek 4 - přečteno ano"])

done_fill = PatternFill("solid", fgColor="D9EAD3")
done_font = Font(name="Calibri", size=9, color="595959")

rule = FormulaRule(
    formula=['AND($A2<>"", $A2<>"☐")'],
    fill=done_fill,
    font=done_font
)
ws.conditional_formatting.add("A2:C100", rule)

wb.save("test_cf.xlsx")
print("Saved test_cf.xlsx successfully")
